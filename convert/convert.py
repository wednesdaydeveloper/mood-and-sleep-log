# -*- coding: utf-8 -*-
"""
過去の Excel セルフモニタリング表をアプリ取込用 CSV に変換する。

入力: convert/{YEAR}セルフモニタリング.xlsx (複数年)
出力: convert/converted.csv (v1.2 アプリ形式 7 列, UTF-8 BOM)

実行: python convert/convert.py
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import Cell


# v1.2 服薬記録の有効キー（眠剤）
SLEEP_AID_DOSES = {0.25, 0.5, 0.75, 1.0, 1.5, 2.0, 2.5, 3.0}

# Excel で「睡眠中」を表す塗りつぶし色（赤）
SLEEP_FILL_RGB = "FFFF0000"

# 「空のシート」と、構造の異なるシートはスキップ
SKIP_SHEETS = {"空のシート", "10"}  # 2025/10 は時間列なしのためスキップ

# CSV 出力時の列順（v1.2）
CSV_HEADER = "date,moodScore,moodTags,memo,sleepIntervals,sleepAid,prnMedication"

# UTF-8 BOM
BOM = "﻿"


# ---------------------------------------------------------------------------
# パース関数
# ---------------------------------------------------------------------------


def parse_date(value: object, year: str) -> str | None:
    """'12/01(月)' のような文字列を 'YYYY-MM-DD' に変換。"""
    if value is None:
        return None
    s = str(value).strip()
    m = re.match(r"^(\d{1,2})/(\d{1,2})", s)
    if not m:
        return None
    mm, dd = int(m.group(1)), int(m.group(2))
    if not (1 <= mm <= 12 and 1 <= dd <= 31):
        return None
    return f"{year}-{mm:02d}-{dd:02d}"


def parse_mood(value: object) -> int:
    """気分列を [-2, +2] の整数に正規化。空欄は 0（普通）扱い。"""
    if value is None or value == "":
        return 0
    if isinstance(value, bool):
        return 0
    if isinstance(value, (int, float)):
        x = float(value)
    else:
        s = str(value).strip().translate(str.maketrans({"＋": "+", "−": "-", "－": "-"}))
        try:
            x = float(s)
        except ValueError:
            return 0
    rounded = int(round(x))
    return max(-2, min(2, rounded))


def parse_sleep_aid(value: object) -> str:
    """眠剤列を 'lunesta-X.X' に変換。1.25 は 1.0 に丸める。
    None/空/'無し' は空欄。
    """
    if value is None:
        return ""
    if isinstance(value, str):
        s = value.strip()
        if s == "" or s == "無し":
            return ""
        try:
            n = float(s)
        except ValueError:
            return ""
    elif isinstance(value, (int, float)):
        n = float(value)
    else:
        return ""
    if n == 1.25:
        n = 1.0  # 仕様: 1.0 に丸め
    if n in SLEEP_AID_DOSES:
        return f"lunesta-{_format_dose(n)}"
    return ""


def parse_prn(value: object) -> str:
    """頓服列の値を変換。
    - None/空/'無し'/0 → 空欄（なし扱い）
    - '有り' → 'lunesta-1.0'
    - 非ゼロ数値 → 'lunesta-{N}'（先頭に 'lunesta-' を付与）
    - その他テキスト → 空欄
    """
    if value is None:
        return ""
    if isinstance(value, str):
        s = value.strip()
        if s == "" or s == "無し":
            return ""
        if s == "有り":
            return "lunesta-1.0"
        try:
            n = float(s)
        except ValueError:
            return ""
    elif isinstance(value, (int, float)):
        n = float(value)
    else:
        return ""
    if n == 0:
        return ""
    return f"lunesta-{_format_dose(n)}"


def _format_dose(n: float) -> str:
    """数値を CSV 用の文字列に。整数値は 'X.0'、小数は通常表記。"""
    if n == int(n):
        return f"{int(n)}.0"
    # 0.25, 0.5, 0.75, 1.5, 2.5 など
    s = f"{n:g}"
    return s


# ---------------------------------------------------------------------------
# 睡眠区間
# ---------------------------------------------------------------------------


def find_hour_columns(header: tuple) -> list[tuple[int, int]]:
    """ヘッダー行から数値の時間列を検出し (col_index_1based, hour) のリストを返す。"""
    cols: list[tuple[int, int]] = []
    for i, h in enumerate(header):
        if isinstance(h, bool):
            continue
        if isinstance(h, (int, float)):
            hour = int(h)
            if 0 <= hour <= 23:
                cols.append((i + 1, hour))  # openpyxl は 1-indexed
    return cols


def hour_to_offset(hour: int) -> int:
    """時刻 (0-23) を 21:00 起点の経過分に変換。"""
    if hour >= 21:
        return (hour - 21) * 60
    return (24 - 21 + hour) * 60


def offset_to_hhmm(offset_min: int) -> str:
    """21:00 起点のオフセット分を HH:MM 形式に。"""
    actual_min = (21 * 60 + offset_min) % (24 * 60)
    h = actual_min // 60
    m = actual_min % 60
    return f"{h:02d}:{m:02d}"


def is_sleep_marker(cell: Cell) -> bool:
    """セルが睡眠を示す塗りつぶし（赤）かを判定。"""
    fill = cell.fill
    if fill is None or fill.patternType != "solid":
        return False
    fg = fill.fgColor
    if fg is None:
        return False
    # rgb / value どちらかが 'FFFF0000'
    rgb = getattr(fg, "rgb", None)
    if rgb == SLEEP_FILL_RGB:
        return True
    val = getattr(fg, "value", None)
    return val == SLEEP_FILL_RGB


def extract_sleep_intervals(ws, row_idx: int, hour_cols: list[tuple[int, int]]) -> str:
    """指定行から睡眠区間を抽出し 'HH:MM-HH:MM,HH:MM-HH:MM' 形式で返す。"""
    # 各バケットの開始オフセット（21:00 起点）を昇順で集める
    filled_offsets: list[int] = []
    for col_idx, hour in hour_cols:
        cell = ws.cell(row=row_idx, column=col_idx)
        if is_sleep_marker(cell):
            filled_offsets.append(hour_to_offset(hour))
    filled_offsets.sort()

    # 連続したバケットを 1 区間に集約（隣接 = 60 分差）
    intervals: list[tuple[int, int]] = []
    i = 0
    while i < len(filled_offsets):
        start = filled_offsets[i]
        end = start + 60
        while i + 1 < len(filled_offsets) and filled_offsets[i + 1] == end:
            end += 60
            i += 1
        intervals.append((start, end))
        i += 1

    return ",".join(f"{offset_to_hhmm(s)}-{offset_to_hhmm(e)}" for s, e in intervals)


# ---------------------------------------------------------------------------
# memo / event 結合
# ---------------------------------------------------------------------------


def build_memo(memo_value: object, event_value: object) -> str:
    """備考とイベントを結合。イベントがあれば末尾に「（イベント：xxx）」を付加。"""
    memo = "" if memo_value is None else str(memo_value).strip()
    event = "" if event_value is None else str(event_value).strip()
    if event == "":
        return memo
    suffix = f"（イベント：{event}）"
    return memo + suffix if memo else suffix


# ---------------------------------------------------------------------------
# 列インデックス検出
# ---------------------------------------------------------------------------


def find_named_columns(header: tuple, names: list[str]) -> dict[str, int | None]:
    """ヘッダー行から指定された列名の 1-indexed 列番号を返す。なければ None。"""
    result: dict[str, int | None] = {n: None for n in names}
    for i, h in enumerate(header):
        if isinstance(h, str):
            for n in names:
                if h == n:
                    result[n] = i + 1
                    break
    return result


# ---------------------------------------------------------------------------
# CSV 出力
# ---------------------------------------------------------------------------


def quote(value: str) -> str:
    """RFC 4180: " で囲み、内部の " は "" にエスケープ。"""
    return '"' + value.replace('"', '""') + '"'


def format_row(date: str, mood: int, memo: str, intervals: str, sleep_aid: str, prn: str) -> str:
    return ",".join([
        date,
        str(mood),
        quote(""),  # moodTags: 常に空
        quote(memo),
        quote(intervals),
        quote(sleep_aid),
        quote(prn),
    ])


# ---------------------------------------------------------------------------
# シート / ファイル処理
# ---------------------------------------------------------------------------


def process_sheet(ws, year: str, sheet_month: int | None) -> tuple[list[str], int, int, int]:
    """1 シートを処理し、CSV 行のリストと統計を返す。
    Returns: (rows, processed_count, empty_skipped, carryover_skipped)

    sheet_month: シート名から推定した月。None ならフィルタしない。
    日付の月が sheet_month と一致しない行はキャリーオーバー（他月シートから取込済み）
    として除外する。
    """
    header_row = next(ws.iter_rows(values_only=True, max_row=1))
    hour_cols = find_hour_columns(header_row)
    named = find_named_columns(header_row, ["気分", "眠剤", "頓服", "備考", "イベント"])

    rows: list[str] = []
    processed = 0
    empty_skipped = 0
    carryover_skipped = 0

    # 2 行目以降を走査
    for row_idx in range(2, ws.max_row + 1):
        date_cell = ws.cell(row=row_idx, column=1).value
        date = parse_date(date_cell, year)
        if date is None:
            empty_skipped += 1
            continue

        # シート名（月）と日付の月が不一致 → キャリーオーバー行としてスキップ
        if sheet_month is not None:
            row_month = int(date.split("-")[1])
            if row_month != sheet_month:
                carryover_skipped += 1
                continue

        def get(col_name: str):
            col = named.get(col_name)
            if col is None:
                return None
            return ws.cell(row=row_idx, column=col).value

        mood = parse_mood(get("気分"))
        memo = build_memo(get("備考"), get("イベント"))
        sleep_aid = parse_sleep_aid(get("眠剤"))
        prn = parse_prn(get("頓服"))
        intervals = extract_sleep_intervals(ws, row_idx, hour_cols)

        rows.append(format_row(date, mood, memo, intervals, sleep_aid, prn))
        processed += 1

    return rows, processed, empty_skipped, carryover_skipped


def process_file(path: Path) -> tuple[list[tuple[str, str]], dict]:
    """1 ファイルを処理し (date, csv_row) のリストとシート別統計を返す。"""
    year_match = re.match(r"(\d{4})", path.name)
    if not year_match:
        print(f"  WARN: ファイル名から年を抽出できません: {path.name}", file=sys.stderr)
        return [], {}
    year = year_match.group(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    all_rows: list[tuple[str, str]] = []
    stats: dict[str, dict] = {}

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            stats[sheet_name] = {"status": "skipped"}
            continue
        ws = wb[sheet_name]
        # シート名が数値なら月として扱い、月不一致のキャリーオーバー行を除外
        try:
            sheet_month = int(sheet_name)
        except ValueError:
            sheet_month = None
        rows, processed, empty_skipped, carryover_skipped = process_sheet(ws, year, sheet_month)
        stats[sheet_name] = {
            "status": "processed",
            "rows": processed,
            "empty_skipped": empty_skipped,
            "carryover_skipped": carryover_skipped,
        }
        # 各行の先頭は date 文字列。並べ替え用に分離
        for r in rows:
            date = r.split(",", 1)[0]
            all_rows.append((date, r))

    return all_rows, stats


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------


def main() -> int:
    convert_dir = Path(__file__).parent
    inputs = sorted(convert_dir.glob("*セルフモニタリング*.xlsx"))
    if not inputs:
        print("入力ファイルが見つかりません", file=sys.stderr)
        return 1

    all_rows: list[tuple[str, str]] = []
    for path in inputs:
        print(f"\n読み込み: {path.name}")
        rows, stats = process_file(path)
        for name, s in stats.items():
            if s["status"] == "skipped":
                print(f"  スキップ: シート '{name}'")
            else:
                extra = ""
                if s.get("carryover_skipped", 0) > 0:
                    extra = f", キャリーオーバー除外 {s['carryover_skipped']}"
                print(f"  シート '{name}': {s['rows']} 行 (空行スキップ {s['empty_skipped']}{extra})")
        all_rows.extend(rows)

    # 日付昇順でソート（ISO 形式なので文字列比較で OK）
    all_rows.sort(key=lambda x: x[0])

    output_path = convert_dir / "converted.csv"
    with open(output_path, "w", encoding="utf-8", newline="") as f:
        f.write(BOM + CSV_HEADER + "\n")
        for _, row in all_rows:
            f.write(row + "\n")

    print(f"\n出力: {output_path} ({len(all_rows)} 行)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
